#Automaticaly recalculate=true
#Single model=false
#������ ����� ��� ���
#������ ����� ��� ���date="01.07.2019"
##################################################
################ Script Section ##################
##################################################

import datetime
import openpyxl
import os
import oily_report as olr

start=2021 #���� ������ ��������
end=2035 #���� ��������� ��������
den = 0.892
start_coll = 3

olr.create_report_dir(path = r'D:\project\Minnib_project\SCHED_PATTERNS\reports\plan_drilling')
print(os.getcwd())

### ���� ��� ������� ����������������� ��������
colls=[(x.to_datetime().year-start) for x in get_all_timesteps ( )]
colls = max(colls)
print(colls)


### ���� ��� ������� ���� ����� ��������
drill_date={}

for m in get_all_models ( ):
	for w in get_all_wells ( ):
		for t in get_all_timesteps ( ):
			 if wstat[m,w,t]==1 or wstat[m,w,t]==2:
			 	drill_date[str(w)] = t.to_datetime()
			 	break
#print(drill_date)


###���� ������ � ������� ��� ������� � excel
temp = openpyxl. Workbook()
ws1 = temp.create_sheet(title='Year_liq_oil')
ws1.merge_cells(start_row=1, start_column=start_coll+1, end_row=1, end_column=colls+start_coll)#.value = '������ �����, �'
ws1.merge_cells(start_row=1, start_column=start_coll+1+colls, end_row=1, end_column=colls*2+start_coll)#.value = '������ ��������, �'
ws1.cell(row=1, column=start_coll+1, value = '������ �����, �')
ws1.cell(row=1, column=start_coll+1+colls,value = '������ ��������, �')

ws1.cell(row=2, column=1, value='WELL') ###����� ������ ����� ������� � ���� excel
ws1.cell(row=2, column=2, value='���� �����') ###����� ���� ����� ������� � ���� excel
ws1.cell(row=2, column=3, value='������') ###����� ���� ����� ������� � ���� excel


for col in range(1, colls+1):
	ws1.cell(row=2, column=start_coll+col, value= start+col) ###��������� ���� ��� �����
	ws1.cell(row=2, column=start_coll+colls+col, value= start+col) ###��������� ���� ��� ��������




### ���� ������� � ������ ������ ������� ������ ����� � �������� �� ���������
row_counter = 2 ##������� �����
for m in get_all_models ( ):
	for w in get_all_wells ( ):
		row_counter += 1
		for t in get_all_timesteps ( ):
			if t.to_datetime().month == 1 and t.to_datetime().year >= start:
				if t.to_datetime().year == start:
					oil_old = float(wopt[m, w, t]/1000)
					liq_old = float(wlpt[m, w, t]/1000)
				else:
					year_oil = float(wopt[m, w, t]/1000)-oil_old
					year_liq = float(wlpt[m, w, t]/1000)-liq_old
					
					column = t.to_datetime().year-start
					ws1.cell(row=row_counter, column=1, value=str(w)) ###����� ������ ������� � ���� excel
					ws1.cell(row=row_counter, column=2, value=drill_date.get(str(w))) ###����� ���� ����� �������� � ������ � ���� excel
					ws1.cell(row=row_counter, column=3, value=str(m)) ###����� ���� ����� �������� � ������ � ���� excel
					ws1.cell(row=row_counter, column=start_coll+column, value= round(year_oil,2)) ###����� ������ ����� ������� � ���� excel
					ws1.cell(row=row_counter, column=start_coll+colls+column, value= round(year_liq,2)) ###����� ������ �������� ������� � ���� excel
					#print (w, row_counter,column+1, round(year_oil,2), round(year_liq,2))
					
					#print (w, t,round(float(wlpt[m, w, t]/1000),1), round(year_oil,2), round(year_liq,2))
					oil_old = float(wopt[m, w, t]/1000)
					liq_old = float(wlpt[m, w, t]/1000)

temp.save('gibrid_years.xlsx')
