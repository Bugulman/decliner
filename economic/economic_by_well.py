#Automaticaly recalculate=true
#Single model=false
#пишите здесь ваш код
#пишите здесь ваш кодdate="01.07.2019"
##################################################
################ Script Section ##################
##################################################

import datetime
import openpyxl
import os
import oily_report as olr
import getpass

start=2008 #дата начала прогноза
end=2038 #дата окончания прогноза
den = 1
start_coll = 3

olr.create_report_dir(path = get_project_folder ( ))
print(os.getcwd())

### Блок для анализа продолжательности прогноза
colls=[(x.to_datetime().year-start) for x in get_all_timesteps ( )]
colls = max(colls)
print(colls)


### Блок для анализа даты ввода скважины
drill_date={}

for m in get_all_models ( ):
	for w in get_all_wells ( ):
		for t in get_all_timesteps ( ):
			 if wstat[m,w,t]==1 or wstat[m,w,t]==2:
			 	drill_date[str(w)] = t.to_datetime()
			 	break
#print(drill_date)


###Блок работы с модулем для импорта в excel
temp = openpyxl. Workbook()
ws1 = temp.create_sheet(title='Year_liq_oil')
ws1.merge_cells(start_row=1, start_column=start_coll+1, end_row=1, end_column=colls+start_coll)#.value = 'Добыча нефти, т'
ws1.merge_cells(start_row=1, start_column=start_coll+1+colls, end_row=1, end_column=colls*2+start_coll)#.value = 'Добыча жидкости, т'
ws1.cell(row=1, column=start_coll+1, value = 'Добыча нефти, т')
ws1.cell(row=1, column=start_coll+1+colls,value = 'Добыча жидкости, т')

ws1.cell(row=2, column=1, value='WELL') ###пишем добычу нефти скважин в файл excel
ws1.cell(row=2, column=2, value='Дата ввода') ###пишем дату ввода скважин в файл excel
ws1.cell(row=2, column=3, value='Модель') ###пишем дату ввода скважин в файл excel


for col in range(1, colls+1):
	ws1.cell(row=2, column=start_coll+col, value= start+col) ###Заполняем годы для нефти
	ws1.cell(row=2, column=start_coll+colls+col, value= start+col) ###Заполняем годы для жидкости




### Блок расчета и вывода данных годовой добыче нефти и жидкости по скважинам
row_counter = 2 ##счетчик строк
for m in get_all_models ( ):
	for w in get_all_wells ( ):
		row_counter += 1
		for t in get_all_timesteps ( ):
			if t.to_datetime().month == 1 and t.to_datetime().year >= start:
				if t.to_datetime().year == start:
					oil_old = float(wopt[m, w, t]/1000)
					liq_old = float(wlpt[m, w, t]/1000)
				else:
					year_oil = float(wopt[m, w, t]/1000)-oil_old
					year_liq = float(wlpt[m, w, t]/1000)-liq_old
					
					column = t.to_datetime().year-start
					ws1.cell(row=row_counter, column=1, value=str(w)) ###пишем номера скважин в файл excel
					ws1.cell(row=row_counter, column=2, value=drill_date.get(str(w))) ###пишем дату ввода скважины в работу в файл excel
					ws1.cell(row=row_counter, column=3, value=str(m)) ###пишем дату ввода скважины в работу в файл excel
					ws1.cell(row=row_counter, column=start_coll+column, value= round(year_oil,2)) ###пишем добычу нефти скважин в файл excel
					ws1.cell(row=row_counter, column=start_coll+colls+column, value= round(year_liq,2)) ###пишем добычу жидкости скважин в файл excel
					#print (w, row_counter,column+1, round(year_oil,2), round(year_liq,2))
					
					#print (w, t,round(float(wlpt[m, w, t]/1000),1), round(year_oil,2), round(year_liq,2))
					oil_old = float(wopt[m, w, t]/1000)
					liq_old = float(wlpt[m, w, t]/1000)

user = getpass.getuser()
date = datetime.datetime.strftime(datetime.datetime.now(), '%d.%m.%y')

temp.save(f'wells_by_year_{user}_{date}.xlsx')
