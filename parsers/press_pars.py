# %%
import os
from support_func import file_runer, convert_excel_dev, convert_csv_KVD, get_noblank_rows, df_resample
from pathlib import Path
from tqdm import tqdm
from sqlalchemy import create_engine
import warnings
import sys
from multiprocessing import Pool
import pandas as pd
import openpyxl
sys.path.append('/home/albert.vafin/Documents/python_proj/decliner/parsers')
os.chdir('/home/albert.vafin/Documents/python_proj/decliner')
warnings.filterwarnings('ignore')
# NOTE: подключение к базе sqllite
engine = create_engine('sqlite:///press.db')

# %%

# Перечень файлов паспортов скважин
# WARN: запускается единожды!
# p = Path(r'/cluster3/public/2023/UNGKM/Data Achimgas/Замеры забойных давлений')

p = Path(r'/cluster3/public/2023/UNGKM/Data Achim Development/info_from_Radmila/telemetry')


# %%
csv_files = file_runer(p, 'CSV', regex_file='GA.*') + \
    file_runer(p, 'csv', regex_file='GA.*')
excel_files = file_runer(p, 'xls') + \
    file_runer(p, 'XLS')

str(excel_files[0])
# %%
files = [
    '/cluster3/public/2023/UNGKM/Data Achim Development/info_from_Radmila/telemetry/2022/Январь/4А16/Январь 4А163.xlsx',
    '/cluster3/public/2023/UNGKM/Data Achim Development/info_from_Radmila/telemetry/2021/Июль/4А11/Июль 4А112.xlsx',
    '/cluster3/public/2023/UNGKM/Data Achim Development/info_from_Radmila/telemetry/2021/Июль/4А17/Июль 4А171.xlsx',
    '/cluster3/public/2023/UNGKM/Data Achim Development/info_from_Radmila/telemetry/2021/Июль/4А17/Июль 4А172.xlsx',
    '/cluster3/public/2023/UNGKM/Data Achim Development/info_from_Radmila/telemetry/2021/Июль/4А17/Июль 4А173.xlsx']

# %%


def convert_excel_dev(excel_file: Path):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    last_row = get_noblank_rows(sheet)
    name = sheet.cell(1, 1).value
    df = pd.read_excel(excel_file, skiprows=5,
                       nrows=last_row-3, usecols=(0, 1, 2, 3),
                       names=["date", "time", "press", "temre"])
    print(df.date.dtypes)
    if df.date.dtypes == '0' or df.date.dtypes == 'object':
        df.date = df.date.astype(str)+' '+df.time.astype(str)
        df.date = pd.to_datetime(df.date, errors='coerce')
    elif df.date.dtypes == '<M8[ns]':
        df.date = pd.to_datetime(df.date, errors='coerce')
    else:
        print(f'не удалось обработать файл{excel_file}')
    df = df[df.date.notnull()]
    df.press = pd.to_numeric(df.press, errors='coerce')
    df.set_index(df.date, inplace=True)
    df = df_resample(df).reset_index()
    df.insert(0, 'well', name)
    df['path'] = str(excel_file)
    return df


# %%
df = convert_excel_dev(Path(excel_files[0]))
df.date
df

# excel_files = excel_files[0:10]
# excel_files
# df[['well', 'date', 'press']]
df.to_csv('test.csv')
# %%

with open('zaboi_excel_dev.txt', 'r') as f:
    file = f.readlines()
    for n in tqdm(file):
        try:
            df = convert_excel_dev(n[:-1])
            df.to_sql('zaboi_excel_dev', con=engine, if_exists='append')
            file.pop(file.index(n))
        except:
            print(f'не считается {n}')

# %%
# NOTE: ИМПОРТ исхдников из экселей в базу данных
with open('zaboi_excel_dev.txt', 'w') as f:
    for n in tqdm(excel_files):
        try:
            df = convert_excel_dev(n)
            df.to_sql('test', con=engine, if_exists='append')
        except:
            f.write(f'{n}\n')
            print(n)

# %%
pbar = tqdm(total=len(excel_files))

def excel_to_base(excel_file):
    with open('zaboi_excel_dev.txt', 'w') as f:
        try:
            df = convert_excel_dev(excel_file)
            df.to_sql('test', con=engine, if_exists='append')
        except:
            f.write(f'{excel_file}\n')
            print(excel_file)
    pbar.update(1)


# %%
import time

start_time = time.time()
with Pool(8) as p:
    p.map(excel_to_base, excel_files)
    print(time.time()-start_time)

# %% NOTE: ИМПОРТ исхдников из csv в базу данных
for n in tqdm(csv_files):
    try:
        df = convert_csv_KVD(n)
        df.to_sql('zaboi_csv_dev', con=engine, if_exists='append')
    except:
        print(f'not{n}')
    # df.to_sql('zaboi_csv', con=engine, if_exists='append')
