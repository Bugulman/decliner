import re
from pathlib import Path
import re
import pandas as pd
from sqlalchemy import create_engine
import openpyxl
import csv
from datetime import datetime
from pathlib import Path


def file_runer(path: Path, file_ext: str, regex_file='.*'):
    '''Функция вытаскивает файлы c расширением
    file_ext из подпапок по пути path. Возможно использование
    регулярок для фильтрации названий файлов'''
    files = []
    counter = 0
    for file in path.glob(f'**/*.{file_ext}*'):
        finder = re.compile(regex_file)
        if finder.search(str(file)) != None:
            try:
                counter += 1
                files.append(file)
            except:
                continue
    print(counter)
    return files


def get_noblank_rows(sheet):
    ''' Ищем последнюю заполненную строку в столбце A'''
    last_row = sheet.max_row
    while sheet.cell(row=last_row, column=1).value is None:
        last_row -= 1
    return last_row


def df_resample(df):
    """Функция преобразует датафрейм в почасовой формат
    с отбраковкой замеров ниже 3 квартиля"""
    resampled = df[df.press > df.press.quantile(0.05)
                   ].press.resample('1H').mean()
    return resampled


def get_date_from_filename(path):
    """функция принимает путь к файлу и формирует дату из названия файла
    в формате yymmdd"""
    date = re.findall('\d{6}', path.parts[-1])
    date = datetime.strptime(date[0], '%y%m%d')
    return date


def get_wells(file_path: str):
    """функция читает заголовок csv файла и заирает оттуда информация по
    имеющимся в файле скважинам"""
    header = []
    with open(file_path, 'r', encoding='cp1251') as var:
        line = csv.reader(var)
        for row in line:
            if len(row) <= 2:
                header.append(row[0].strip())
            else:
                break
    return header


def convert_datetime_to_date(date):
    return date.strftime('%d.%m.%Y') if type(date) == datetime else date


def convert_excel_KVD(excel_file: Path):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    last_row = get_noblank_rows(sheet)
    name = sheet.cell(1, 1).value
    date = sheet.cell(2, 2).value
    date = convert_datetime_to_date(date)
    date = datetime.strptime(str(date), '%d.%m.%Y')
    df = pd.read_excel(excel_file, skiprows=3,
                       nrows=last_row-3, usecols=(0, 1),
                       names=["date", "press"])
    df.date = str(date) + ' ' + df.date.astype('str')
    df.date = pd.to_datetime(df.date)
    df.set_index(df.date, inplace=True)
    df = df_resample(df).reset_index()
    df.insert(0, 'well', name)
    df['path'] = str(excel_file)
    # print(df)
    return df


def convert_csv_KVD(file_path: Path):
    header = get_wells(str(file_path))
    date = get_date_from_filename(file_path)
    full_frame = pd.read_csv(file_path, skiprows=len(
        header)+2, engine='python', encoding='cp1251')
    # print(header)
    df = pd.DataFrame()
    for seq in range(len(header)):
        # print(header[seq])
        subframe = full_frame.iloc[:, [0, 2*seq+1, 2*seq+2]]
        # print(subframe.head())
        subframe.columns = ['date', 'press', 'temp']
        # print(subframe.head())
        subframe.date = pd.to_timedelta(subframe.date, errors='coerce')
        subframe = subframe[~subframe.date.isnull()]
        subframe.date = subframe.date+date
        # print(subframe)
        # subframe.index = subframe.date
        subframe.press = pd.to_numeric(subframe.press)
        subframe.temp = pd.to_numeric(subframe.temp)
        subframe.set_index(subframe.date, inplace=True)
        subframe = df_resample(subframe).reset_index()
        # subframe.to_csv('testt.csv')
        # subframe = subframe[['press', 'temp']].resample('H').mean()
        subframe['well'] = header[seq]
        subframe['path'] = str(file_path)
        # subframe.to_sql('press', con=engine, if_exists='append')
        df = pd.concat([df, subframe], ignore_index=True)
    return df
