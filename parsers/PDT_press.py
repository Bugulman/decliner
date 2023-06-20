# %%
from pprint import pprint
import time
from pathlib import Path
import re
import pandas as pd
from sqlalchemy import create_engine
from tqdm import tqdm
import openpyxl
import csv
from datetime import datetime, time
import warnings
import matplotlib.pyplot as plt
warnings.filterwarnings('ignore')
p = Path.cwd()
p
# NOTE: подключение к базе sqllite
engine = create_engine('sqlite:///press.db')

# %%

# WARN: запускается единожды!
p = Path(r'/cluster3/public/2023/UNGKM/Data Achimgas/ГДИ, ГКИ')
print(p)


def file_runer(path: Path, file_ext: str, regex_file='.*'):
    '''Функция вытаскивает файлы c расширением
    file_ext из подпапок по пути path. Возможно использование
    регулярок для фильтрации названий файлов'''
    with open('press.txt', 'w+') as var:
        files = []
        counter = 0
        for file in path.glob(f'**/*.{file_ext}*'):
            finder = re.compile(regex_file)
            if finder.search(str(file)) != None:
                try:
                    var.write(str(file)+'\n')
                    # print(file)
                    counter += 1
                    files.append(file)
                except:
                    continue
    print(counter)
    return files


def get_noblank_rows(sheet):
    ''' Ищем последнюю заполненную строку в столбце A'''
    last_row = sheet.max_row
    while sheet.cell(row=last_row, column=1).value is None:
        last_row -= 1
    return last_row


def df_resample(df):
    """Функция преобразует датафрейм в почасовой формат 
    с отбраковкой замеров ниже 3 квартиля"""
    resampled = df[df.press > df.press.quantile(0.05)
                   ].press.resample('1H').mean()
    return resampled


def get_date_from_filename(path):
    """функция принимает путь к файлу и формирует дату из названия файла
    в формате yymmdd"""
    date = re.findall('\d{6}', path.parts[-1])
    date = datetime.strptime(date[0], '%y%m%d')
    return date


def get_wells(file_path: str):
    """функция читает заголовок csv файла и заирает оттуда информация по 
    имеющимся в файле скважинам"""
    header = []
    with open(file_path, 'r', encoding='cp1251') as var:
        line = csv.reader(var)
        for row in line:
            if len(row) <= 2:
                header.append(row[0].strip())
            else:
                break
    return header


def convert_excel_KVD(excel_file: Path):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    last_row = get_noblank_rows(sheet)
    name = sheet.cell(1, 1).value
    date = sheet.cell(2, 2).value
    date = datetime.strptime(date, '%d.%m.%Y')
    df = pd.read_excel(excel_file, skiprows=3,
                       nrows=last_row-3, usecols=(0, 1),
                       names=["date", "press"])
    df.date = str(date) + ' ' + df.date.astype('str')
    df.date = pd.to_datetime(df.date)
    df.set_index(df.date, inplace=True)
    df = df_resample(df).reset_index()
    df.insert(0, 'well', name)
    df['path'] = str(excel_file)
    # print(df)
    return df


def convert_csv_KVD(file_path: Path):
    header = get_wells(str(file_path))
    date = get_date_from_filename(file_path)
    full_frame = pd.read_csv(file_path, skiprows=len(
        header)+2, engine='python', encoding='cp1251')
    # print(header)
    df = pd.DataFrame()
    for seq in range(len(header)):
        print(header[seq])
        subframe = full_frame.iloc[:, [0, 2*seq+1, 2*seq+2]]
        # print(subframe.head())
        subframe.columns = ['date', 'press', 'temp']
        # print(subframe.head())
        subframe.date = pd.to_timedelta(subframe.date, errors='coerce')
        subframe = subframe[~subframe.date.isnull()]
        subframe.date = subframe.date+date
        # print(subframe)
        # subframe.index = subframe.date
        subframe.press = pd.to_numeric(subframe.press)
        subframe.temp = pd.to_numeric(subframe.temp)
        subframe.set_index(subframe.date, inplace=True)
        subframe = df_resample(subframe).reset_index()
        # subframe.to_csv('testt.csv')
        # subframe = subframe[['press', 'temp']].resample('H').mean()
        subframe['well'] = header[seq]
        subframe['path'] = str(file_path)
        # subframe.to_sql('press', con=engine, if_exists='append')
        df = pd.concat([df, subframe], ignore_index=True)
    return df

    # print(subframe)
    # else:
    # print(row[0], row[1], row[2])
    # subframe.info()


# %%
# NOTE: АНАЛИЗ и создание списков файлов с различными расширениями
csv_files = file_runer('CSV', regex_file='GA.*')
excel_files = file_runer('xls')
txt_files = file_runer('txt')

pprint(csv_files[0:10])
# %%
# NOTE: ИМПОРТ исхдников из экселей в базу данных
for n in tqdm(excel_files):
    try:
        df = convert_excel_KVD(n)
        df.to_sql('KVD_excel', con=engine, if_exists='append')
    except:
        print(n)


# %%
# NOTE: ИМПОРТ исхдников из csv в базу данных
for n in tqdm(csv_files):
    try:
        df = convert_csv_KVD(n)
        df.to_sql('KVD_csv', con=engine, if_exists='append')
    except:
        print(n)


#  %%BUG:УПРАЖНЕНИЯ чтобы все работало как надо
df = convert_excel_KVD(excel_files[33])
df = convert_csv_KVD(csv_files[0])
df.columns
df.date.dt.day
df.set_index(df.date, inplace=True)  # .date.dt.date
df.press.groupby(pd.Grouper(freq='1H')).mean()


# %%
df.info()
df.set_index(df.date, inplace=True)
resampled = df[df.press > df.press.quantile(0.05)
               ].press.resample('1H').mean()
resampled.plot()
plt.show()
df[df.press == 0]
# %%

fr = pd.read_sql('KVD_excel', con=engine)
fr['day'] = fr.kvd_time.dt.date
fr.set_index(fr.kvd_time, inplace=True)  # .date.dt.date
fr[['well', 'press']].groupby(pd.Grouper(key='well', freq='1D')).mean()
fr.kvd_time.value_counts()
# %% BUG: заплаточка
for w in fr.well.unique():
    sub = fr.loc[fr.well == w]
    resampled_df = df_resample(sub).reset_index()
    resampled_df['well'] = w.split(' ')[-1].replace('-', '')
    print(resampled_df)
    resampled_df.to_sql('resampled_KVD', con=engine, if_exists='append')
# %%
plt.show()
sub.press.resample('1H').mean()
