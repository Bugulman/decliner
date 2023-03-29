#!/usr/bin/env python3
# %%
import time
from pathlib import Path
import re
import pandas as pd
import xlwings as xw
import openpyxl
import xlrd
from sqlalchemy import create_engine
from tqdm import tqdm
import numpy as np

p = Path.cwd()

print(p)

# %%
# работает долго! НЕ ЗАПУСКАЙ!
for file in p.glob('**/*.xls*'):
    if 'рапорт' in str(file) and '~' not in str(file):
        print(file)

# %%
engine = create_engine(
    'postgresql://test:test@localhost:5434/ungkm')
# %%

# обрабатвает файлы для получения списка экселей с названием листов
files = []
sheets = []
prod_files = {}
for file in p.glob('**/*.xls*'):
    if 'рапорт' in str(file) and '~' not in str(file):
        try:
            book = openpyxl.load_workbook(file)
            sheet = book.sheetnames
            files.append(str(file))
            sheets.append(sheet)
        except:
            book = xlrd.open_workbook(file)
            sheet = book.sheet_names()
            files.append(str(file))
            sheets.append(sheet)
prod_files['files'] = files
prod_files['sheets'] = sheets
print(prod_files)

# %%

well_list = ['действующий фонд',
             'бездействующий фонд', 'ликвидированный фонд']

# %%


def find_excel_row(sheet, to_find, column):
    """TODO: Docstring for find_excel_row.
    :returns: TODO
    """
    try:
        start_cell = sheet[column].value.index(float(to_find))
        start_cell += 1
        table = sheet.range((start_cell, 1), (start_cell, 36))
    except Exception as e:
        regex_str = re.compile(to_find)
        column_value = [x for x in sheet['A:A'].value if x != None]
        column_value = list(map(str, column_value))
        temp = [regex_str.findall(x)
                for x in column_value if regex_str.search(x) != None]
        print(temp)
        start_cell = sheet[column].value.index(temp[0][0])
        start_cell += 1
        table = sheet.range((start_cell, 1), (start_cell, 36))
    return table


def parse_excel_table(file, reg_ex, column):
    '''функция считывает таблицу добычи из переданного файла, при
    этом автоматически расширяя в файле эксель вверх и в стороны'''
    # print(file)
    with xw.App(visible=False) as app:
        # app = xw.App()
        book = xw.Book(file)
        name = book.sheet_names[0]
        sheet = book.sheets[name]
        try:
            table = find_excel_row(sheet, reg_ex, column)
            df = table.options(pd.DataFrame, expand='down').value
            df.replace(to_replace='-', value=np.nan, inplace=True)
            df.replace(to_replace=' -', value=np.nan, inplace=True)
            df.columns = range(0, df.shape[1])
            # book.save(file)
            book.close()
            # book.app.kill()
            return df[df[2].notnull()]
        except Exception as e:
            print(
                f'Ошибка поиска значения {reg_ex} в файле {file}')
            return None


def add_metadata(df, extra_data):
    for k, v in extra_data.items():
        df[k] = v

# %%


files = pd.read_excel(p.parent.joinpath('mer_files.xlsx'),
                      sheet_name='Sheet1')

files = files[files['to_take'] > 0]
files.drop('to_take', axis=1, inplace=True)
# %%

for k, v in files.iloc[0].items():
    print(k, v)

# %%

work_wells = pd.DataFrame()
for file in tqdm(files.iterrows()):
    content = file[1]['Path']
    try:
        df2 = parse_excel_table(content, '[Дд]ействующий.*', 'A:A')
        add_metadata(df2, file[1])
        work_wells = pd.concat([work_wells, df2])
    except Exception as e:
        df2 = parse_excel_table(content, '1', 'A:A')
        add_metadata(df2, file[1])
        work_wells = pd.concat([work_wells, df2])

work_wells['status'] = 'work'  # .to_excel('remove.xlsx')

# %%

new_wells = pd.DataFrame()
for file in tqdm(files.iterrows()):
    content = file[1]['Path']
    try:
        df1 = parse_excel_table(
            content, '.*строительстве.*', 'A:A')
        add_metadata(df1, file[1])
        new_wells = pd.concat([new_wells, df1])
    except Exception as e:
        continue

new_wells['status'] = 'new'  # .shape

# %%
work_wells

# %%

work_wells = pd.concat([work_wells, new_wells])
col_names = ['kust', 'well', 'obj', 'pust', 'pvh', 'tust', 'tvh',
             'gas_av', 'gas_prod', 'gas_burn', 'gas_tot', 'sep_gas_av',
             'sep_gas_prod', 'sep_gas_burn', 'sep_gas_tot', 'dry_gas_prod',
             'dry_gas_burn', 'dry_gas_tot', 'calc_coeff', 'SCHU',
             'cond_content', 'cond_prod', 'cond_use', 'cond_lost',
             'cond_res', 'cond_burn', 'cond_tot', 'unstable_con_prod',
             'wat_av', 'wat_prod', 'time_work', 'time_stay', 'time_calend',
             'stay_cod', 'coeff_uch', 'path', 'file', 'sheets', 'date', 'status']

work_wells.columns = col_names

# %%

work_wells.to_sql('prod', con=engine, if_exists='append')
# %%

# WARNING: часть для тестирования

some = '1'
float(some)

book = xw.Book(p.joinpath
               (r'2007\Сентябрь 2007\рапорт форма 1 за сентябрь скорректированный.xls'))

book = xw.Book(p.joinpath
               (r'2021\09\Эксплуатационный рапорт\Эксплуатационный рапорт 04-10-2021 10-28.xlsx'))
name = book.sheet_names[0]
sheet = book.sheets[name]
sheet

book.close()
book.app.kill()


df = find_excel_row(sheet, '1', 'A:A')
df = df.options(pd.DataFrame, expand='down').value
df = parse_excel_table(p.joinpath
                       (r'2007\Сентябрь 2007\рапорт форма 1 за сентябрь скорректированный.xls'),
                       '1', 'A:A')

df1 = parse_excel_table(p.joinpath
                        (r'2017\09\Эксплуатационный рапорт\Ежемесячный рапорт 03-10-2017 10-12.xlsx'),
                        '.*строительстве.*', 'A:A')
df1.isinstance(pd.DataFrame())

find_excel_row(sheet, '[Дд]ействующий.*', 'A:A')

start_cell = sheet['A:A'].value.index(float(some))

start_cell += 1

a = 2 if 22 < 0 else 3
a

start_cell

table = sheet.range((start_cell, 1), (start_cell, 36))

table

df = find_excel_row(sheet, '[Бб]ездействующий.*', 'A:A')
df = find_excel_row(sheet, '[Дд]ействующий.*', 'A:A')
df = df.options(pd.DataFrame, expand='down').value
df.columns = range(0, df.shape[1])
df

regex_str = re.compile('[Бб]ездействующий.*')
column_value = [x for x in sheet['A:A'].value if x != None]
column_value = list(map(str, column_value))
column_value

temp = [regex_str.findall(x)
        for x in column_value if regex_str.search(x) != None]
temp

start_cell = sheet['A:A'].value.index(temp[0][0])
start_cell

# %%

book = openpyxl.load_workbook(p.joinpath
                              (r'2007\Октябрь 2007\рапорт форма 1 за октябрь скорректированный.xls'))

# %%
book = xlrd.open_workbook(p.joinpath
                          (r'2007\Октябрь 2007\рапорт форма 1 за октябрь скорректированный.xls'))
# %%

mylist = [1, 2, 3, 4, 5]

for i in tqdm(mylist):
    time.sleep(1)

# %%
